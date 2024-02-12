from kivymd.app import MDApp
from kivymd.uix.screen import Screen
from kivy.uix.textinput import TextInput
from kivy.uix.screenmanager import ScreenManager
from kivymd.uix.datatables import MDDataTable
from kivymd.icon_definitions import md_icons
from kivymd.uix.button import MDIconButton
from kivy.metrics import dp
from kivy.uix.button import Button
from kivy.graphics import Color, RoundedRectangle
import pandas as pd
import sqlite3
from sort_data import sort_price
from kivy.clock import Clock
import webbrowser
from kivymd.toast import toast
from openpyxl import load_workbook
from kivy.core.window import Window
import openpyxl
import os
from kivy.uix.label import Label
from kivymd.uix.label import MDLabel
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
import random
username = ''
from kivy.uix.popup import Popup
primary_key = ''
guest = ''
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, NoTransition
from kivy.uix.button import Button
#Used to check if items are in the cart
checked_items = []
#Storing item index
store_ids = []
#Cart counter
cart_count = 0
#Stores the data
data = []
text1 = ""
screen_one_bool = True
screen_three_bool = False
screen_four_bool = False
screen_five_bool = True
screen_six_bool = False
screen_seven_bool = False
sorted_high_data = []
sorted_low_data = []
from sort_data import sort_price2
new_data = []
from saving_data import save_data3
'''
Creates the default screen
'''
class ScreenOne(Screen):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        Window.bind(on_request_close=lambda *args: True)

        # Create the data table
        self.data_table1 = MDDataTable(
            #background_color_selected_cell="e4514f",
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,

            column_data=[
                ("ID.", dp(25)),
                ("Website", dp(25)),
                ("Title", dp(150)),
                ("Price", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=data,


        )



        self.data_table1.bind(on_check_press=self.on_check_press)
        self.add_widget(self.data_table1)
        # Creates icon button
        refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(refresh_button)


        refresh_button.bind(on_press=self.refresh_main_data)
        # Creates sorting high prices button
        high_button = Button(
            text="Sort High",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.10, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        high_button.bind(on_press=self.sort_high)
        self.add_widget(high_button)
        # Creates sorting low prices button
        low_button = Button(
            text="Sort Low",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.35, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        low_button.bind(on_press=self.sort_low)
        self.add_widget(low_button)

        # Create the check cart button
        check_cart = Button(
            text="Check Cart",
            size_hint=(0.3, 0.1),
            pos_hint={'center_x': 0.5, 'center_y': 0.1},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        check_cart.bind(on_press=self.go_to_screen_two)
        self.add_widget(check_cart)

        # Create the save button
        save_button = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color = (1, 1, 1, 1),
            border = (0, 0, 0, 5),

        )
        save_button.bind(on_press=self.save_file)
        self.add_widget(save_button)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )
        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)

        # Create input text field
        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Enter/Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_input)

        # Create a button for saving the entered item ID
        save_id_button = Button(
            text="Save ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )


        save_id_button.bind(on_press=self.save_item_id)
        self.add_widget(save_id_button)
        # Create item cart
        cart_layout = BoxLayout(
            size_hint=(None, None),
            size=(dp(120), dp(48)),
            pos_hint={'center_x': 0.5, 'center_y': 0.9},
            spacing=dp(10)
        )
        cart_icon = MDIconButton(
            icon="cart",
            theme_text_color="Custom",

        )
        self.cart_label = MDLabel(
            text="0",
            size_hint=(None, None),
            pos_hint={'center_x': 0.455, 'center_y': 0.94},
            size=(dp(32), dp(32)),
            halign="center",
            valign="center",
            font_style="Body2",
            theme_text_color="Custom",

        )
        cart_layout.add_widget(cart_icon)
        self.add_widget(cart_layout)
        self.add_widget(self.cart_label)
    # Refresh cart
    def refresh_cart(self):
        # Get the current value of the cart and update the label
        #current_cart_value = cart_count  # replace this with the function that gets the current cart value
        self.cart_label.text = f"{cart_count}"




    # Refreshes the data
    def refresh_main_data(self, button):
        self.remove_widget(self.data_table1)
        self.data_table1 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Website", dp(25)),
                ("Title", dp(150)),
                ("Price", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=data,

        )
        self.add_widget(self.data_table1)
    # Saves item ID
    def save_item_id(self, instance):
        global cart_count
        global checked_items
        global store_ids
        # Text input
        try:
            item_id = int(self.item_id_input.text)

        # Checks if the input is a positive number and less than data length
            if item_id < len(data) and item_id >= 0:
                # If the item id is stored remove it from the list and update cart
                if item_id in store_ids:
                    # Gets the Item ID
                    stored_index = store_ids.index(item_id)
                    del store_ids[stored_index]
                    del checked_items[stored_index]
                    cart_count -= 1
                    self.item_id_input.text = ""
                # Add item ID to the cart and update the cart
                else:
                    # Remove item from list if it is unchecked
                    # print(current_row[0])
                    store_ids.append(item_id)
                    checked_items.append(data[item_id])
                    #print("Price: ",data[item_id][3])
                    #print(checked_items)
                    cart_count += 1
                    self.item_id_input.text = ""
                self.cart_label.text = f"{cart_count}"
                # print(checked_items)
            else:
                show_popup(1, 1)
        except:
            show_popup(1, 1)

    def on_check_press(self, instance_table, current_row):
        global cart_count
        global checked_items
        global store_ids
        # print(instance_table)
        # checkbox = current_row.children[0].children[0]
        index = current_row[0]
        print(current_row[1])
        print(instance_table)
        # print(index)

        if index in store_ids:
            # index = store_instance.index(instance_table)
            stored_index = store_ids.index(index)
            del store_ids[stored_index]
            del checked_items[stored_index]
            cart_count -= 1
        else:
            # Remove item from list if it is unchecked
            print(current_row[0])
            store_ids.append(index)
            checked_items.append(current_row)
            cart_count += 1
        self.cart_label.text = f"Cart: {cart_count}"
        # print(checked_items)
        # Saves the display file
    '''
    Saves data
    '''
    def save_file(self, button):
        # connecting a path
        if guest == False:
            workbook = openpyxl.load_workbook(f'{text1}.xlsx')
            # Select the active worksheet
            worksheet = workbook.active

            # Get the user's Downloads folder path
            try:
                # For Windows
                downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                print("winodw")
            except KeyError:
                try:
                    # For macOS/iOS
                    downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    print("Mac")
                except KeyError:
                    # For Linux and other platforms
                    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
                    print("other")

            # Save the file to the Downloads folder with a filename
            filename = os.path.join(downloads_folder, f'{text1}.xlsx')

            # Use try-except block to catch the `PermissionError` exception
            try:
                workbook.save(filename)
                show_popup(0, 0)
            except PermissionError as e:
                show_popup(0, 6)
        else:
            show_popup(0, 3)




    #Quit method
    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids


        conn = sqlite3.connect('userdata.db')
        cur = conn.cursor()

        # Select the primary key of the user with the given username
        cur.execute("SELECT rowid FROM userdata WHERE username = ?", (username,))
        result = cur.fetchone()

        # Close the connection to the database


        # Check if a matching user was found
        if result is not None:
            # Retrieve the primary key value
            primary_key = result[0]
            print("Primary key for user", username, "is", primary_key)
        else:
            print("User", username, "not found")
            primary_key = 0
        for item in checked_items:
            cur.execute("""
            INSERT INTO items (user_id, Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (primary_key, item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8], item[9], item[10]))

        # Commit the changes

        conn.commit()
        conn.close()
        if text1 == "":
            pass
        else:
            try:
                cart_count = 0
                checked_items.clear()
                store_ids.clear()
                data.clear()
                os.remove(f'{text1}.xlsx')
                os.remove(f'{text1}_high.xlsx')
                os.remove(f'{text1}_low.xlsx')
                MyApp.get_running_app().stop()
            except:
                pass

    # Accesses the low price excel file and screen
    def sort_low(self, button):
        if guest == False:
            global screen_one_bool
            global screen_four_bool
            screen_one_bool = False
            screen_four_bool = True
            screen_one = self.manager.get_screen('screen_four')
            screen_one.refresh_cart()
            self.manager.current = 'screen_four'
        else:
            show_popup(0, 3)

    # Accesses the high price excel file and screen
    def sort_high(self, button):
        if guest == False:
            global screen_three_bool
            global screen_one_bool
            screen_three_bool = True
            screen_one_bool = False
            screen_one = self.manager.get_screen('screen_three')
            screen_one.refresh_cart()
            self.manager.current = 'screen_three'
        else:
            show_popup(0, 3)

    # Goes to cart
    def go_to_screen_two(self, *args):
        screen_two = self.manager.get_screen('screen_two')
        screen_two.refresh_data(checked_items)
        self.manager.current = 'screen_two'

'''
Screen two is used to display the cart
'''
class ScreenTwo(Screen):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Creates check cart button
        check_cart = Button(
            text="Check data",
            size_hint=(0.3, 0.1),
            pos_hint={'center_x': 0.5, 'center_y': 0.1},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        check_cart.bind(on_press=self.go_to_screen_one)
        self.add_widget(check_cart)
        #Data table
        self.data_table2 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Website", dp(25)),
                ("Title", dp(150)),
                ("Price", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=checked_items,

        )

        self.add_widget(self.data_table2)
        # Creates save selected items button
        save_select_items = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        save_select_items.bind(on_press=self.save_selected_items)
        self.add_widget(save_select_items)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )
        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)
        # Refreshes the datatable
        refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(refresh_button)
        refresh_button.bind(on_press=self.refresh_data)

        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},


        )
        self.add_widget(self.item_id_input)

        # Create a button for saving the entered item ID
        self.remove_id_button = Button(
            text="Remove ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        self.remove_id_button.bind(on_press=self.remove_item_id)
        self.remove_id_button.bind(on_release=self.refresh_data)

        self.add_widget(self.remove_id_button)
        # Enter the item ID for link
        self.item_id_link_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Enter ID for link",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.24, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_link_input)

        # Create a button for saving the entered item ID
        self.enter_link = Button(
            text="Get Link",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.34, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        self.enter_link.bind(on_press=self.open_link)
        #self.enter_link.bind(on_release=self.refresh_data)

        self.add_widget(self.enter_link)

        self.data_table2.bind(on_check_press=self.on_row_press)

    def on_row_press(self, instance_table, instance_row):
        print(instance_row, instance_table)
    '''
    Used to open the link
    '''
    def open_link(self, button):
        try:
            item_id = int(self.item_id_link_input.text)
            stored_index = store_ids.index(item_id)
            try:
                webbrowser.open(checked_items[stored_index][10])
                self.item_id_link_input.text = ''
            except:
                show_popup(0,5)

        except:
            show_popup(0,4)



    # Goes back to the previously called screen
    def go_to_screen_one(self, *args):
        if screen_one_bool:
            screen_one = self.manager.get_screen('screen_one')
            screen_one.refresh_cart()
            self.manager.current = 'screen_one'
        elif screen_three_bool:
            screen_three = self.manager.get_screen('screen_three')
            screen_three.refresh_cart()
            self.manager.current = 'screen_three'
        elif screen_four_bool:
            screen_three = self.manager.get_screen('screen_four')
            screen_three.refresh_cart()
            self.manager.current = 'screen_four'
    # Refreshing the table when an item is removed
    def refresh_data(self, button):
        self.remove_widget(self.data_table2)
        self.data_table2 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Website", dp(25)),
                ("Title", dp(150)),
                ("Price", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=checked_items,

        )
        self.add_widget(self.data_table2)
    '''
    Remove ID function 
    '''
    def remove_item_id(self, instance):
        global cart_count
        global checked_items
        global store_ids
        # Text input
        try:
            item_id = int(self.item_id_input.text)

            # Checks if the input is a positive number and less than data length
            if item_id < len(data) and item_id >= 0:
                # If the item id is stored remove it from the list and update cart
                if item_id in store_ids:
                    # index = store_instance.index(instance_table)
                    stored_index = store_ids.index(item_id)
                    del store_ids[stored_index]
                    del checked_items[stored_index]
                    cart_count -= 1
                    self.item_id_input.text = ""

                # Add item ID to the cart and update the cart
                else:
                    print("cant only remove items")
            else:
                show_popup(1, 1)
        except:
            show_popup(1, 1)

    # Used to download the items in the cart
    def save_selected_items(self, button):
        if guest == False:
            # Checks if items are in the cart
            if len(checked_items) == 0:
                show_popup(1, 2)
            else:
                selected_rows = []
                for row_index in checked_items:

                    selected_row = {"ID": row_index[0],"Website":row_index[1], "Title": row_index[2], "Price": row_index[3],
                                    "Stock": row_index[4],
                                    "Rating": row_index[5], "Review": row_index[6], "Description": row_index[7],
                                    "Color": row_index[8],
                                    "Size": row_index[9], "Link": row_index[10]}

                    selected_rows.append(selected_row)

                # Create a pandas DataFrame from the selected rows
                selected_df = pd.DataFrame(selected_rows)

                # Random number code for saving data
                rand_num = random.randint(1, 10 ** 100)
                # Save the DataFrame to an Excel file
                filename = f"{text1}selected_items{rand_num}.csv"

                # Get the user's Downloads folder path
                try:
                    # For Windows
                    downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                except KeyError:
                    try:
                        # For macOS/iOS
                        downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    except KeyError:
                        # For Linux and other platforms
                        downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")

                # Create the full file path
                filepath = os.path.join(downloads_folder, filename)

                # Save the DataFrame to the file path
                selected_df.to_csv(filepath, index=False)
                show_popup(0, 0)
        else:
            show_popup(0, 3)


    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids


        conn = sqlite3.connect('userdata.db')
        cur = conn.cursor()

        # Select the primary key of the user with the given username
        cur.execute("SELECT rowid FROM userdata WHERE username = ?", (username,))
        result = cur.fetchone()

        # Close the connection to the database


        # Check if a matching user was found
        if result is not None:
            # Retrieve the primary key value
            primary_key = result[0]
            print("Primary key for user", username, "is", primary_key)
        else:
            print("User", username, "not found")

        for item in checked_items:
            cur.execute("""
            INSERT INTO items (user_id, Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (primary_key, item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8], item[9], item[10]))

        # Commit the changes

        conn.commit()
        conn.close()
        if text1 == "":
            pass
        else:
            try:
                cart_count = 0
                checked_items.clear()
                store_ids.clear()
                data.clear()
                os.remove(f'{text1}.xlsx')
                os.remove(f'{text1}_high.xlsx')
                os.remove(f'{text1}_low.xlsx')
                MyApp.get_running_app().stop()
            except:
                pass

class ScreenThree(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        Window.bind(on_request_close=lambda *args: True)

        # Create the data table
        self.data_table3 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),

            ],
            row_data=sorted_high_data,

        )



        self.data_table3.bind(on_check_press=self.on_check_press)
        self.add_widget(self.data_table3)

        # Create the save selected button
        check_cart = Button(
            text="Check Cart",
            size_hint=(0.3, 0.1),
            pos_hint={'center_x': 0.5, 'center_y': 0.1},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        check_cart.bind(on_press=self.go_to_screen_two)
        self.add_widget(check_cart)

        # Create the save button
        save_button = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        save_button.bind(on_press=self.save_file)
        self.add_widget(save_button)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )

        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)
        default_button = Button(
            text="Default",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.10, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        default_button.bind(on_press=self.sort_default)

        self.add_widget(default_button)
        # Create input text field
        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Enter/Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_input)

        # Create a button for low prices
        low_button = Button(
            text="Sort Low",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.35, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        low_button.bind(on_press=self.sort_low)
        self.add_widget(low_button)
        # Creates Save id button
        save_id_button = Button(
            text="Save ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        save_id_button.bind(on_press=self.save_item_id)
        self.add_widget(save_id_button)
        # Create refresh button
        refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(refresh_button)
        refresh_button.bind(on_press=self.refresh_main_data)
        # Create item cart
        cart_layout = BoxLayout(
            size_hint=(None, None),
            size=(dp(120), dp(48)),
            pos_hint={'center_x': 0.5, 'center_y': 0.9},
            spacing=dp(10)
        )
        cart_icon = MDIconButton(
            icon="cart",
            theme_text_color="Custom",

        )
        # Creates cart label
        self.cart_label = MDLabel(
            text="0",
            size_hint=(None, None),
            pos_hint={'center_x': 0.455, 'center_y': 0.94},
            size=(dp(32), dp(32)),
            halign="center",
            valign="center",
            font_style="Body2",
            theme_text_color="Custom",

        )
        cart_layout.add_widget(cart_icon)
        self.add_widget(cart_layout)
        self.add_widget(self.cart_label)


    def go_to_screen_two(self, *args):
        screen_two = self.manager.get_screen('screen_two')
        screen_two.refresh_data(checked_items)
        self.manager.current = 'screen_two'

    def refresh_cart(self):
        # Get the current value of the cart and update the label

        self.cart_label.text = f"{cart_count}"

    def refresh_main_data(self, button):
        self.remove_widget(self.data_table3)
        self.data_table3 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=sorted_high_data,

        )
        self.add_widget(self.data_table3)

    def sort_low(self, button):
        global screen_three_bool
        global screen_four_bool
        screen_three_bool = False
        screen_four_bool = True
        screen_one = self.manager.get_screen('screen_four')
        screen_one.refresh_cart()
        self.manager.current = 'screen_four'

    def save_item_id(self, instance):
        global cart_count
        global checked_items
        global store_ids
        # Text input
        try:
            item_id = int(self.item_id_input.text)
        # Checks if the input is a positive number and less than data length
            if item_id < len(data) and item_id >= 0:
                # If the item id is stored remove it from the list and update cart
                if item_id in store_ids:
                    # index = store_instance.index(instance_table)
                    stored_index = store_ids.index(item_id)
                    del store_ids[stored_index]
                    del checked_items[stored_index]
                    cart_count -= 1
                    self.item_id_input.text = ""
                # Add item ID to the cart and update the cart
                else:
                    # Remove item from list if it is unchecked

                    store_ids.append(item_id)
                    checked_items.append(data[item_id])

                    cart_count += 1
                    self.item_id_input.text = ""
                self.cart_label.text = f"{cart_count}"

            else:
                show_popup(1, 1)
        except:
            show_popup(1, 1)


    def on_check_press(self, instance_table, current_row):
        global cart_count
        global checked_items
        global store_ids

        index = current_row[0]
        print(current_row[1])
        print(instance_table)
        # print(index)

        if index in store_ids:
            stored_index = store_ids.index(index)
            del store_ids[stored_index]
            del checked_items[stored_index]
            cart_count -= 1
        else:
            # Remove item from list if it is unchecked
            print(current_row[0])
            store_ids.append(index)
            checked_items.append(current_row)
            cart_count += 1
        self.cart_label.text = f"Cart: {cart_count}"

    # Saves the display file
    def save_file(self, button):
        # connecting a path
        if guest == False:
            workbook = openpyxl.load_workbook(f'{text1}_high.xlsx')
            # Select the active worksheet
            worksheet = workbook.active

            # Get the user's Downloads folder path
            try:
                # For Windows
                downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                print("winodw")
            except KeyError:
                try:
                    # For macOS/iOS
                    downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    print("Mac")
                except KeyError:
                    # For Linux and other platforms
                    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
                    print("other")

            # Save the file to the Downloads folder with a filename
            filename = os.path.join(downloads_folder, f'{text1}_high.xlsx')
            try:
                workbook.save(filename)
                show_popup(0, 0)
            except PermissionError as e:
                show_popup(0, 6)
        else:
            show_popup(0, 3)


    def sort_default(self, button):
        global screen_one_bool
        global screen_three_bool
        screen_one_bool = True
        screen_three_bool = False
        screen_one = self.manager.get_screen('screen_one')
        screen_one.refresh_cart()
        self.manager.current = 'screen_one'

    #Quit method
    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids


        conn = sqlite3.connect('userdata.db')
        cur = conn.cursor()

        # Select the primary key of the user with the given username
        cur.execute("SELECT rowid FROM userdata WHERE username = ?", (username,))
        result = cur.fetchone()

        # Close the connection to the database


        # Check if a matching user was found
        if result is not None:
            # Retrieve the primary key value
            primary_key = result[0]
            print("Primary key for user", username, "is", primary_key)
        else:
            print("User", username, "not found")

        for item in checked_items:
            cur.execute("""
            INSERT INTO items (user_id, Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (primary_key, item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8], item[9], item[10]))

        # Commit the changes


        conn.commit()
        conn.close()
        if text1 == "":
            pass
        else:
            try:
                cart_count = 0
                checked_items.clear()
                store_ids.clear()
                data.clear()
                os.remove(f'{text1}.xlsx')
                os.remove(f'{text1}_high.xlsx')
                os.remove(f'{text1}_low.xlsx')
                MyApp.get_running_app().stop()
            except:
                pass


class ScreenFour(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        Window.bind(on_request_close=lambda *args: True)

        # Create the data table
        self.data_table4 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),

            ],
            row_data=sorted_low_data,

        )

        # data_table.row_data[0]['check'] = False

        self.data_table4.bind(on_check_press=self.on_check_press)
        self.add_widget(self.data_table4)

        # Create the save selected button
        check_cart = Button(
            text="Check Cart",
            size_hint=(0.3, 0.1),
            pos_hint={'center_x': 0.5, 'center_y': 0.1},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        check_cart.bind(on_press=self.go_to_screen_two)
        self.add_widget(check_cart)
        low_button = Button(
            text="Sort High",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.35, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        low_button.bind(on_press=self.sort_high)
        self.add_widget(low_button)
        # Create the save button
        save_button = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        save_button.bind(on_press=self.save_file)
        self.add_widget(save_button)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )

        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)
        default_button = Button(
            text="Default",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.10, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        default_button.bind(on_press=self.sort_default)

        self.add_widget(default_button)
        # Create input text field
        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Enter/Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_input)

        # Create a button for saving the entered item ID
        save_id_button = Button(
            text="Save ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        save_id_button.bind(on_press=self.save_item_id)
        self.add_widget(save_id_button)
        # Create refresh button
        refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(refresh_button)
        refresh_button.bind(on_press=self.refresh_main_data)
        # Create item cart
        cart_layout = BoxLayout(
            size_hint=(None, None),
            size=(dp(120), dp(48)),
            pos_hint={'center_x': 0.5, 'center_y': 0.9},
            spacing=dp(10)
        )
        cart_icon = MDIconButton(
            icon="cart",
            theme_text_color="Custom",

        )
        self.cart_label = MDLabel(
            text="0",
            size_hint=(None, None),
            pos_hint={'center_x': 0.455, 'center_y': 0.94},
            size=(dp(32), dp(32)),
            halign="center",
            valign="center",
            font_style="Body2",
            theme_text_color="Custom",

        )
        cart_layout.add_widget(cart_icon)
        self.add_widget(cart_layout)
        self.add_widget(self.cart_label)


    def go_to_screen_two(self, *args):
        screen_two = self.manager.get_screen('screen_two')
        screen_two.refresh_data(checked_items)
        self.manager.current = 'screen_two'

    def refresh_cart(self):
        # Get the current value of the cart and update the label

        self.cart_label.text = f"{cart_count}"

    def refresh_main_data(self, button):
        self.remove_widget(self.data_table4)
        self.data_table4 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=sorted_low_data,

        )
        self.add_widget(self.data_table4)


    def save_item_id(self, instance):
        global cart_count
        global checked_items
        global store_ids
        # Text input
        try:
            item_id = int(self.item_id_input.text)


        # Checks if the input is a positive number and less than data length
            if item_id < len(data) and item_id >= 0:
                # If the item id is stored remove it from the list and update cart
                if item_id in store_ids:
                    # index = store_instance.index(instance_table)
                    stored_index = store_ids.index(item_id)
                    del store_ids[stored_index]
                    del checked_items[stored_index]
                    cart_count -= 1
                    self.item_id_input.text = ""
                # Add item ID to the cart and update the cart
                else:
                    # Remove item from list if it is unchecked

                    store_ids.append(item_id)
                    checked_items.append(data[item_id])

                    cart_count += 1
                    self.item_id_input.text = ""
                self.cart_label.text = f"{cart_count}"

            else:
                show_popup(1, 1)
        except:
            show_popup(1, 1)


    def on_check_press(self, instance_table, current_row):
        global cart_count
        global checked_items
        global store_ids

        index = current_row[0]
        print(current_row[1])
        print(instance_table)
        # print(index)

        if index in store_ids:

            stored_index = store_ids.index(index)
            del store_ids[stored_index]
            del checked_items[stored_index]
            cart_count -= 1
        else:
            # Remove item from list if it is unchecked
            print(current_row[0])
            store_ids.append(index)
            checked_items.append(current_row)
            cart_count += 1
        self.cart_label.text = f"Cart: {cart_count}"

        # Saves the display file

    def save_file(self, button):
        # connecting a path
        if guest == False:
            workbook = openpyxl.load_workbook(f'{text1}_low.xlsx')


            # Get the user's Downloads folder path
            try:
                # For Windows
                downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                print("winodw")
            except KeyError:
                try:
                    # For macOS/iOS
                    downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    print("Mac")
                except KeyError:
                    # For Linux and other platforms
                    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
                    print("other")

            # Save the file to the Downloads folder with a filename
            filename = os.path.join(downloads_folder, f'{text1}_low.xlsx')
            try:
                workbook.save(filename)
                show_popup(0, 0)
            except PermissionError as e:
                show_popup(0, 6)
        else:
            show_popup(0, 3)


    def sort_default(self, button):
        global screen_one_bool
        global screen_four_bool
        screen_one_bool = True
        screen_four_bool = False
        screen_one = self.manager.get_screen('screen_one')
        screen_one.refresh_cart()
        self.manager.current = 'screen_one'

    def sort_high(self, button):
        global screen_three_bool
        global screen_four_bool
        screen_three_bool = True
        screen_four_bool = False
        screen_one = self.manager.get_screen('screen_three')
        screen_one.refresh_cart()
        self.manager.current = 'screen_three'

    #Quit method
    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids


        conn = sqlite3.connect('userdata.db')
        cur = conn.cursor()

        # Select the primary key of the user with the given username
        cur.execute("SELECT rowid FROM userdata WHERE username = ?", (username,))
        result = cur.fetchone()

        # Close the connection to the database


        # Check if a matching user was found
        if result is not None:
            # Retrieve the primary key value
            primary_key = result[0]
            print("Primary key for user", username, "is", primary_key)
        else:
            print("User", username, "not found")

        for item in checked_items:
            cur.execute("""
            INSERT INTO items (user_id, Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (primary_key, item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8], item[9], item[10]))

        # Commit the changes

            # Commit the changes
        conn.commit()
        conn.close()
        if text1 == "":
            pass
        else:
            try:
                cart_count = 0
                checked_items.clear()
                store_ids.clear()
                data.clear()
                os.remove(f'{text1}.xlsx')
                os.remove(f'{text1}_high.xlsx')
                os.remove(f'{text1}_low.xlsx')
                MyApp.get_running_app().stop()
            except:
                pass



class ScreenManagement(ScreenManager):
    pass

def show_popup(index_t, index_s):
    sentences = ["Data Saved", "Invalid number", "Must select items to use this button", "Guest users cannot save data", "ID is not in the cart", "Sorry, product link is not valid", "Please close all excel sheets"]
    titles = ["Success", "Error"]
    content = Label(text=sentences[index_s])
    popup = Popup(title=titles[index_t], content=content, size_hint=(None, None), size=(400, 400))
    popup.open()
    # close popup after 3 seconds
    Clock.schedule_once(popup.dismiss, 3)

class MyApp(MDApp):
    def build(self):
        screen_manager = ScreenManagement()
        self.theme_cls.theme_style = "Light"
        screen_manager.add_widget(ScreenOne(name='screen_one'))
        screen_manager.add_widget(ScreenTwo(name='screen_two'))
        screen_manager.add_widget(ScreenThree(name='screen_three'))
        screen_manager.add_widget(ScreenFour(name='screen_four'))
        return screen_manager

    def go_to_screen_two(self, instance):
        self.root.current = 'screen_two'

'''
Used for built in scrapers
text: string
guest1: boolean
sorts data and displays the data
'''
def run_data(text, guest1, user):
    global data
    global text1
    global guest
    global sorted_high_data
    global sorted_low_data
    global username
    username = user
    guest = guest1
    print(guest)
    text1 = text
    # Read Excel file
    df = pd.read_excel(f'{text}.xlsx')

    # Stores the data
    data = []

    for row in df.itertuples(index=False):
        data.append((row[0], row[1],row[2], row[3], row[4], row[5], row[6], row[7], row[8],  row[9], row[10]))


    sorted_high_data, sorted_low_data = sort_price(text, data)
    MyApp().run()

'''
Used for extracted data in master data file
text: string
Used to sort prices
'''
def run_data2(text, user):
    global data
    global text1
    global guest
    global sorted_high_data
    global sorted_low_data
    global username
    username = user
    guest = False
    text1 = f'filtered_data{text}'
    df = pd.read_excel(f'{text1}.xlsx')


    # Stores the data
    data = []
    for row in df.itertuples(index=True):
        data.append((row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10]))


    sorted_high_data, sorted_low_data = sort_price(text1, data)
    MyApp().run()


'''
def sort_price(text,data):
    data_2 = []
    for key in data:
        if ' - ' in key[3]:
            lower, upper = key[3].split(' - ')
            average = (float(lower.replace('£', '')) + float(upper.replace('£', ''))) / 2
            average_float = round(float(average), 2)
            data_2.append((key[0], average_float, key[1], key[2],  key[4], key[5], key[6], key[7], key[8], key[9]))
        elif 'to' in key[3]:
            lower, upper = key[3].split(' to ')

            average = (float(lower.replace('£', '')) + float(upper.replace('£', ''))) / 2
            average_float = round(float(average), 2)
            data_2.append((key[0],average_float, key[1], key[2],  key[4], key[5], key[6], key[7], key[8], key[9]))
        else:
            remove_sign = key[3].replace('£', '').replace(',', '')
            try:
                remove_sign = float(remove_sign)
            except:
                remove_sign = float(-1.0)
            data_2.append((key[0], remove_sign, key[1], key[2], key[4], key[5], key[6], key[7], key[8], key[9]))
    sorted_low_data = sorted(data_2, key=lambda x: x[1], reverse=False)
    sorted_high_data = sorted(data_2, key=lambda x: x[1], reverse=True)

    df = pd.DataFrame(sorted_high_data, columns=["Index","Price", "Website", "Description", "Stock", "Review", "Color", "Rating", "Size", "Link"])
    df2 = pd.DataFrame(sorted_low_data, columns=["Index","Price", "Website", "Description", "Stock", "Review", "Color", "Rating", "Size", "Link"])
    filename = f"{text}_high.xlsx"
    filename2 = f"{text}_low.xlsx"
    # Create an Excel writer object


    # Write the DataFrame to an Excel sheet named "Sheet1"


    # Save the Excel file
    df.to_excel(filename, index=False)
    df2.to_excel(filename2, index=False)
'''
class ScreenFive(Screen):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        Window.bind(on_request_close=lambda *args: True)

        # Create the data table
        self.data_table1 = MDDataTable(
            #background_color_selected_cell="e4514f",
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,

            column_data=[
                ("ID.", dp(25)),
                ("Website", dp(25)),
                ("Title", dp(150)),
                ("Price", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=data,


        )



        self.data_table1.bind(on_check_press=self.on_check_press)
        self.add_widget(self.data_table1)
        # Creates icon button
        refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(refresh_button)


        refresh_button.bind(on_press=self.refresh_main_data)
        # Creates sorting high prices button
        high_button = Button(
            text="Sort High",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.10, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        high_button.bind(on_press=self.sort_high)
        self.add_widget(high_button)
        # Creates sorting low prices button
        low_button = Button(
            text="Sort Low",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.35, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        low_button.bind(on_press=self.sort_low)
        self.add_widget(low_button)

        # Create the check cart button


        # Create the save button
        save_button = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color = (1, 1, 1, 1),
            border = (0, 0, 0, 5),

        )
        save_button.bind(on_press=self.save_file)
        self.add_widget(save_button)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )
        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)

        # Create input text field
        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_input)

        # Create a button for saving the entered item ID
        self.remove_id_button = Button(
            text="Remove ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )


        self.remove_id_button.bind(on_press=self.save_item_id)

        self.add_widget(self.remove_id_button)
        # Create item cart
        cart_layout = BoxLayout(
            size_hint=(None, None),
            size=(dp(120), dp(48)),
            pos_hint={'center_x': 0.5, 'center_y': 0.9},
            spacing=dp(10)
        )
        cart_icon = MDIconButton(
            icon="cart",
            theme_text_color="Custom",

        )
        self.cart_label = MDLabel(
            text="0",
            size_hint=(None, None),
            pos_hint={'center_x': 0.455, 'center_y': 0.94},
            size=(dp(32), dp(32)),
            halign="center",
            valign="center",
            font_style="Body2",
            theme_text_color="Custom",

        )
        cart_layout.add_widget(cart_icon)
        self.add_widget(cart_layout)
        self.add_widget(self.cart_label)
    # Refresh cart





    # Refreshes the data
    def refresh_main_data(self, button):
        self.remove_widget(self.data_table1)
        self.data_table1 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Website", dp(25)),
                ("Title", dp(150)),
                ("Price", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=data,

        )
        self.add_widget(self.data_table1)
        self.cart_label.text = f"{len(data)}"

    '''
    Removes ID from the cart and reorgizes the data strcutre 
    '''
    def save_item_id(self, instance):

        # Text input
        try:
            item_id = int(self.item_id_input.text)


        # Checks if the input is a positive number and less than data length

            if item_id in store_ids:
                # Gets the Item ID
                print(data[item_id][2])
                conn = sqlite3.connect('userdata.db')

                cursor = conn.cursor()


                try:
                    # Deletes the specific item from the database
                    cursor.execute(
                        "DELETE FROM items WHERE id=(SELECT id FROM items WHERE user_id=? AND Website=? AND Title=? AND Price=? AND Stock=? AND Rating=? AND Review=? AND Description=? AND Color=? AND Size=? AND Link=? LIMIT 1)",
                        (primary_key, data[item_id][1], data[item_id][2], data[item_id][3], data[item_id][4],
                         data[item_id][5], data[item_id][6], data[item_id][7], data[item_id][8], data[item_id][9],
                         data[item_id][10]))
                    conn.commit()
                    idx = store_ids.index(item_id)
                    del store_ids[idx]
                    del data[idx]

                    del self.data_table1.row_data[idx]
                    # Selects the data from the database
                    cursor.execute(
                        "SELECT Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link FROM items WHERE user_id=?",
                        (primary_key,))
                    items = cursor.fetchall()
                    data.clear()
                    for idx in range(len(items)):
                        store_ids.append(idx)
                        print(store_ids)
                        data.append((idx, items[idx][0], items[idx][1], items[idx][2], items[idx][3], items[idx][4],
                                     items[idx][5], items[idx][6], items[idx][7], items[idx][8], items[idx][9]))


                    self.refresh_cart(data)
                    self.item_id_input.text = ""
                except:
                    pass

            else:
                show_popup(1, 1)
        except:
            show_popup(1, 1)

    def on_check_press(self, instance_table, current_row):
        global cart_count
        global checked_items
        global store_ids

        index = current_row[0]
        print(current_row[1])
        print(instance_table)


        if index in store_ids:

            stored_index = store_ids.index(index)
            del store_ids[stored_index]
            del checked_items[stored_index]
            cart_count -= 1
        else:
            # Remove item from list if it is unchecked
            print(current_row[0])
            store_ids.append(index)
            checked_items.append(current_row)
            cart_count += 1
        self.cart_label.text = f"Cart: {cart_count}"

        # Saves the display file
    '''
    Saves data
    '''
    def save_file(self, button):
        # connecting a path
        if guest == False:
            save_data3(data, "cart")
            workbook = openpyxl.load_workbook(f'cart.xlsx')
            # Select the active worksheet
            worksheet = workbook.active

            # Get the user's Downloads folder path
            try:
                # For Windows
                downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                print("winodw")
            except KeyError:
                try:
                    # For macOS/iOS
                    downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    print("Mac")
                except KeyError:
                    # For Linux and other platforms
                    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
                    print("other")

            # Save the file to the Downloads folder with a filename
            filename = os.path.join(downloads_folder, f'cart.xlsx')

            # Use try-except block to catch the `PermissionError` exception
            try:
                workbook.save(filename)
                show_popup(0, 0)
            except PermissionError as e:
                show_popup(0, 6)
        else:
            show_popup(0, 3)




    #Quit method
    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids
        if text1 == "":
            pass
        else:
            try:
                os.remove(f'cart.xlsx')
            except:
                pass
            try:
                os.remove(f'carthigh.xlsx')
            except:
                pass
            try:
                os.remove(f'cartlow.xlsx')
            except:
                pass

            cart_count = 0
            checked_items.clear()
            store_ids.clear()
            data.clear()
            MyApp.get_running_app().stop()

    #Accesses the low price excel file and screen
    def sort_low(self, button):
        #_, low = sort_price2(data)
        global screen_five_bool
        global screen_seven_bool
        screen_five_bool = False
        screen_seven_bool = True
        #screen_three = self.manager.get_screen('screen_three')
        _,new_data = sort_price2(data)
        screen_one = self.manager.get_screen('screen_seven')
        screen_one.refresh_cart(new_data)
        self.manager.current = 'screen_seven'
    #Accesses the high price excel file and screen
    def sort_high(self, button):
        global new_data
        global screen_five_bool
        global screen_six_bool
        new_data, _ = sort_price2(data)
        print(new_data)
        screen_five_bool = False
        screen_six_bool = True
        # screen_three = self.manager.get_screen('screen_three')
        screen_one = self.manager.get_screen('screen_six')
        screen_one.refresh_cart(new_data)
        self.manager.current = 'screen_six'

    def refresh_cart(self, new_data):
        self.data_table1.row_data = new_data
        self.cart_label.text = f"{len(data)}"




class ScreenSix(Screen):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        Window.bind(on_request_close=lambda *args: True)

        # Create the data table
        self.data_table1 = MDDataTable(
            #background_color_selected_cell="e4514f",
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,

            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=[],


        )



        self.data_table1.bind(on_check_press=self.on_check_press)
        self.add_widget(self.data_table1)
        # Creates icon button
        refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(refresh_button)


        refresh_button.bind(on_press=self.refresh_main_data)
        # Creates sorting high prices button
        high_button = Button(
            text="Default",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.10, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        high_button.bind(on_press=self.sort_default)
        self.add_widget(high_button)
        # Creates sorting low prices button
        low_button = Button(
            text="Sort Low",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.35, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        low_button.bind(on_press=self.sort_low)
        self.add_widget(low_button)

        # Create the check cart button


        # Create the save button
        save_button = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color = (1, 1, 1, 1),
            border = (0, 0, 0, 5),

        )
        save_button.bind(on_press=self.save_file)
        self.add_widget(save_button)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )
        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)

        # Create input text field
        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_input)

        # Create a button for saving the entered item ID
        self.remove_id_button = Button(
            text="Remove ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )


        self.remove_id_button.bind(on_press=self.save_item_id)
        #self.remove_id_button.bind(on_release=self.refresh_main_data)
        self.add_widget(self.remove_id_button)
        # Create item cart
        cart_layout = BoxLayout(
            size_hint=(None, None),
            size=(dp(120), dp(48)),
            pos_hint={'center_x': 0.5, 'center_y': 0.9},
            spacing=dp(10)
        )
        cart_icon = MDIconButton(
            icon="cart",
            theme_text_color="Custom",

        )
        self.cart_label = MDLabel(
            text="0",
            size_hint=(None, None),
            pos_hint={'center_x': 0.455, 'center_y': 0.94},
            size=(dp(32), dp(32)),
            halign="center",
            valign="center",
            font_style="Body2",
            theme_text_color="Custom",

        )
        cart_layout.add_widget(cart_icon)
        self.add_widget(cart_layout)
        self.add_widget(self.cart_label)
    # Refresh cart
    def refresh_cart(self, new_data):
        self.data_table1.row_data = new_data
        self.cart_label.text = f"{len(new_data)}"




        # Refreshes the data
    def refresh_main_data(self, button):
        self.remove_widget(self.data_table1)
        self.data_table1 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=data,

        )
        self.add_widget(self.data_table1)
        self.cart_label.text = f"{len(data)}"
    # Saves item ID
    def save_item_id(self, instance):
        stored_data = []
        item_id = int(self.item_id_input.text)
        try:
            for d in self.data_table1.row_data:
                stored_data.append(d[0])
        except:
            pass

        if item_id in stored_data:
            # Gets the Item ID
            conn = sqlite3.connect('userdata.db')
            cursor = conn.cursor()

            try:
                cursor.execute(
                    "DELETE FROM items WHERE id=(SELECT id FROM items WHERE user_id=? AND Website=? AND Title=? AND Price=? AND Stock=? AND Rating=? AND Review=? AND Description=? AND Color=? AND Size=? AND Link=? LIMIT 1)",
                    (primary_key, data[item_id][1], data[item_id][2], data[item_id][3], data[item_id][4],
                     data[item_id][5], data[item_id][6], data[item_id][7], data[item_id][8], data[item_id][9],
                     data[item_id][10]))
                conn.commit()
                idx = stored_data.index(item_id)
                idx2 = store_ids.index(item_id)
                del self.data_table1.row_data[idx]
                del data[idx2]
                cursor.execute(
                    "SELECT Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link FROM items WHERE user_id=?",
                    (primary_key,))
                items = cursor.fetchall()
                data.clear()
                for idx in range(len(items)):
                    store_ids.append(idx)
                    print(store_ids)
                    data.append((idx, items[idx][0], items[idx][1], items[idx][2], items[idx][3], items[idx][4],
                                 items[idx][5], items[idx][6], items[idx][7], items[idx][8], items[idx][9]))

                #new_data = self.data_table1.row_data
                high, _ = sort_price2(data)
                self.refresh_cart(high)

                # del checked_items[stored_index]
                # cart_count -= 1
                self.item_id_input.text = ""
            except:
                print('error')
                self.refresh_cart(self.data_table1.row_data)
                pass

        else:
            show_popup(1, 1)

        #print(stored_data)

    def on_check_press(self, instance_table, current_row):
        global cart_count
        global checked_items
        global store_ids

        index = current_row[0]
        print(current_row[1])
        print(instance_table)


        if index in store_ids:

            stored_index = store_ids.index(index)
            del store_ids[stored_index]
            del checked_items[stored_index]
            cart_count -= 1
        else:
            # Remove item from list if it is unchecked
            print(current_row[0])
            store_ids.append(index)
            checked_items.append(current_row)
            cart_count += 1
        self.cart_label.text = f"Cart: {cart_count}"
        # print(checked_items)
        # Saves the display file
    '''
    Saves data
    '''
    def save_file(self, button):
        # connecting a path
        if guest == False:
            save_data3(self.data_table1.row_data, "carthigh")
            workbook = openpyxl.load_workbook(f'carthigh.xlsx')
            # Select the active worksheet
            worksheet = workbook.active

            # Get the user's Downloads folder path
            try:
                # For Windows
                downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                print("winodw")
            except KeyError:
                try:
                    # For macOS/iOS
                    downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    print("Mac")
                except KeyError:
                    # For Linux and other platforms
                    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
                    print("other")

            # Save the file to the Downloads folder with a filename
            filename = os.path.join(downloads_folder, f'carthigh.xlsx')

            # Use try-except block to catch the `PermissionError` exception
            try:
                workbook.save(filename)
                show_popup(0, 0)
            except PermissionError as e:
                show_popup(0, 6)
        else:
            show_popup(0, 3)




    #Quit method
    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids
        if text1 == "":
            pass
        else:
            try:
                os.remove(f'cart.xlsx')
            except:
                pass
            try:
                os.remove(f'carthigh.xlsx')
            except:
                pass
            try:
                os.remove(f'cartlow.xlsx')
            except:
                pass

            cart_count = 0
            checked_items.clear()
            store_ids.clear()
            data.clear()
            MyApp.get_running_app().stop()

    # Accesses the low price excel file and screen
    def sort_low(self, button):
        _, low = sort_price2(data)
        global screen_five_bool
        global screen_seven_bool
        screen_five_bool = False
        screen_seven_bool = True

        screen_one = self.manager.get_screen('screen_seven')
        screen_one.refresh_cart(low)
        self.manager.current = 'screen_seven'
    # Accesses the high price excel file and screen
    def sort_default(self, button):
        global screen_five_bool
        global screen_seven_bool
        screen_five_bool = False
        screen_seven_bool = True

        screen_one = self.manager.get_screen('screen_five')
        screen_one.refresh_cart(data)
        self.manager.current = 'screen_five'


class ScreenSeven(Screen):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        Window.bind(on_request_close=lambda *args: True)

        # Create the data table
        self.data_table1 = MDDataTable(
            #background_color_selected_cell="e4514f",
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,

            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=[],


        )



        self.data_table1.bind(on_check_press=self.on_check_press)
        self.add_widget(self.data_table1)
        # Creates icon button
        self.refresh_button = MDIconButton(
            icon="refresh",
            pos_hint={'center_x': 0.06, 'center_y': 0.9},
        )
        self.add_widget(self.refresh_button)


        self.refresh_button.bind(on_press=self.refresh_main_data)
        # Creates sorting high prices button
        high_button = Button(
            text="Default",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.10, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        high_button.bind(on_press=self.sort_default)
        self.add_widget(high_button)
        # Creates sorting low prices button
        low_button = Button(
            text="Sort High",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.35, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )
        low_button.bind(on_press=self.sort_high)
        self.add_widget(low_button)

        # Create the check cart button


        # Create the save button
        save_button = Button(
            text="Save",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.9, 'center_y': 0.1},
            background_normal='',
            background_color=(0.1, 0.5, 0.1, 1.0),
            color = (1, 1, 1, 1),
            border = (0, 0, 0, 5),

        )
        save_button.bind(on_press=self.save_file)
        self.add_widget(save_button)

        # Create the quit button
        quit_button = Button(
            text="Leave",
            size_hint=(0.1, 0.1),
            pos_hint={'center_x': 0.1, 'center_y': 0.1},
            background_normal='',
            background_color=(0.8, 0.2, 0.2, 1.0),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),

        )
        quit_button.bind(on_press=self.quit)
        self.add_widget(quit_button)

        # Create input text field
        self.item_id_input = TextInput(
            size_hint=(0.4, None),
            height=dp(30),
            hint_text="Remove Item ID",
            multiline=False,
            input_filter="int",
            pos_hint={'center_x': 0.74, 'center_y': 0.83},
        )
        self.add_widget(self.item_id_input)

        # Create a button for saving the entered item ID
        self.remove_id_button = Button(
            text="Remove ID",
            size_hint=(0.2, None),
            height=dp(30),
            pos_hint={'center_x': 0.84, 'center_y': 0.83},
            background_normal='',
            background_color=(0, 0.349, 0.686, 1),
            color=(1, 1, 1, 1),
            border=(0, 0, 0, 5),
        )


        self.remove_id_button.bind(on_press=self.save_item_id)

        self.add_widget(self.remove_id_button)
        # Create item cart
        cart_layout = BoxLayout(
            size_hint=(None, None),
            size=(dp(120), dp(48)),
            pos_hint={'center_x': 0.5, 'center_y': 0.9},
            spacing=dp(10)
        )
        cart_icon = MDIconButton(
            icon="cart",
            theme_text_color="Custom",

        )
        self.cart_label = MDLabel(
            text="0",
            size_hint=(None, None),
            pos_hint={'center_x': 0.455, 'center_y': 0.94},
            size=(dp(32), dp(32)),
            halign="center",
            valign="center",
            font_style="Body2",
            theme_text_color="Custom",

        )
        cart_layout.add_widget(cart_icon)
        self.add_widget(cart_layout)
        self.add_widget(self.cart_label)
    # Refresh cart
    def refresh_cart(self, new_data):
        self.data_table1.row_data = new_data
        self.cart_label.text = f"{len(new_data)}"

        # Refreshes the data
    def refresh_main_data(self, button):
        self.remove_widget(self.data_table1)
        self.data_table1 = MDDataTable(
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            size_hint=(1.0, 0.6),
            check=False,
            use_pagination=True,
            rows_num=10,
            column_data=[
                ("ID.", dp(25)),
                ("Price", dp(25)),
                ("Title", dp(150)),
                ("Website", dp(25)),
                ("Stock", dp(25)),
                ("Rating", dp(25)),
                ("Review", dp(25)),
                ("Description", dp(350)),
                ("Color", dp(150)),
                ("Size", dp(150)),
                ("Link", dp(200)),
            ],
            row_data=data,

        )
        self.add_widget(self.data_table1)
        self.cart_label.text = f"{len(data)}"
    # Saves item ID
    def save_item_id(self, instance):
        stored_data = []
        item_id = int(self.item_id_input.text)
        try:
            for d in self.data_table1.row_data:
                stored_data.append(d[0])
        except:
            pass


        if item_id in stored_data:
            # Gets the Item ID
            conn = sqlite3.connect('userdata.db')

            cursor = conn.cursor()

            try:
                cursor.execute(
                    "DELETE FROM items WHERE id=(SELECT id FROM items WHERE user_id=? AND Website=? AND Title=? AND Price=? AND Stock=? AND Rating=? AND Review=? AND Description=? AND Color=? AND Size=? AND Link=? LIMIT 1)",
                    (primary_key, data[item_id][1], data[item_id][2], data[item_id][3], data[item_id][4],
                     data[item_id][5], data[item_id][6], data[item_id][7], data[item_id][8], data[item_id][9],
                     data[item_id][10]))
                conn.commit()
                idx = stored_data.index(item_id)
                idx2 = store_ids.index(item_id)
                del self.data_table1.row_data[idx]
                del data[idx2]
                cursor.execute(
                    "SELECT Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link FROM items WHERE user_id=?",
                    (primary_key,))
                items = cursor.fetchall()
                data.clear()
                for idx in range(len(items)):
                    store_ids.append(idx)
                    print(store_ids)
                    data.append((idx, items[idx][0], items[idx][1], items[idx][2], items[idx][3], items[idx][4],
                                 items[idx][5], items[idx][6], items[idx][7], items[idx][8], items[idx][9]))


                _,low = sort_price2(data)
                self.refresh_cart(low)

                self.item_id_input.text = ""
            except:
                print('error')
                pass

        else:
            show_popup(1, 1)

        print(stored_data)


    def on_check_press(self, instance_table, current_row):
        global cart_count
        global checked_items
        global store_ids

        index = current_row[0]
        print(current_row[1])
        print(instance_table)
        # print(index)

        if index in store_ids:

            stored_index = store_ids.index(index)
            del store_ids[stored_index]
            del checked_items[stored_index]
            cart_count -= 1
        else:
            # Remove item from list if it is unchecked
            print(current_row[0])
            store_ids.append(index)
            checked_items.append(current_row)
            cart_count += 1
        self.cart_label.text = f"Cart: {cart_count}"
        # print(checked_items)
        # Saves the display file
    '''
    Saves data
    '''
    def save_file(self, button):
        # connecting a path
        if guest == False:
            save_data3(self.data_table1.row_data, "cartlow")
            workbook = openpyxl.load_workbook(f'cartlow.xlsx')
            # Select the active worksheet
            worksheet = workbook.active

            # Get the user's Downloads folder path
            try:
                # For Windows
                downloads_folder = os.path.join(os.environ["HOMEPATH"], "Downloads")
                print("winodw")
            except KeyError:
                try:
                    # For macOS/iOS
                    downloads_folder = os.path.join(os.environ["HOME"], "Downloads")
                    print("Mac")
                except KeyError:
                    # For Linux and other platforms
                    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
                    print("other")

            # Save the file to the Downloads folder with a filename
            filename = os.path.join(downloads_folder, f'cartlow.xlsx')

            # Use try-except block to catch the `PermissionError` exception
            try:
                workbook.save(filename)
                show_popup(0, 0)
            except PermissionError as e:
                show_popup(0, 6)
        else:
            show_popup(0, 3)




    #Quit method
    def quit(self, button):
        global cart_count
        global checked_items
        global store_ids
        if text1 == "":
            pass
        else:
            try:
                os.remove(f'cart.xlsx')
            except:
                pass
            try:
                os.remove(f'carthigh.xlsx')
            except:
                pass
            try:
                os.remove(f'cartlow.xlsx')
            except:
                pass

            cart_count = 0
            checked_items.clear()
            store_ids.clear()
            data.clear()
            MyApp.get_running_app().stop()


    # Accesses the low price excel file and screen
    def sort_high(self, button):
        high, _ = sort_price2(data)
        global screen_six_bool
        global screen_seven_bool
        screen_six_bool = True
        screen_seven_bool = False
        #screen_three = self.manager.get_screen('screen_three')
        screen_one = self.manager.get_screen('screen_six')
        screen_one.refresh_cart(high)
        self.manager.current = 'screen_six'

    # Accesses the high price excel file and screen
    def sort_default(self, button):
        global screen_five_bool
        global screen_seven_bool
        screen_five_bool = False
        screen_seven_bool = True
        # screen_three = self.manager.get_screen('screen_three')
        screen_one = self.manager.get_screen('screen_five')
        screen_one.refresh_cart(data)
        self.manager.current = 'screen_five'


class MyApp2(MDApp):
    def build(self):
        screen_manager = ScreenManagement()
        self.theme_cls.theme_style = "Light"
        screen_manager.add_widget(ScreenFive(name='screen_five'))
        screen_manager.add_widget(ScreenSix(name='screen_six'))
        screen_manager.add_widget(ScreenSeven(name='screen_seven'))
        #screen_manager.add_widget(ScreenSeven(name='screen_seven'))
        return screen_manager

def run_data3(text, user):
    global data
    global text1
    global guest
    global sorted_high_data
    global sorted_low_data
    global store_ids
    global username
    username = user
    global primary_key
    guest = False
    text1 = f'filtered_data{text}'
    conn = sqlite3.connect('userdata.db')

    cursor = conn.cursor()
    cursor.execute("SELECT rowid FROM userdata WHERE username = ?", (username,))
    result = cursor.fetchone()

    if result is not None:
        # Retrieve the primary key value
        primary_key = result[0]
        print(primary_key)
        print("Primary key for user", username, "is", primary_key)
    else:
        primary_key = 0
        print("User", username, "not found")

    cursor.execute("SELECT Website, Title, Price, Stock, Rating, Review, Description, Color, Size, Link FROM items WHERE user_id=?", (primary_key,))
    items = cursor.fetchall()
    for idx in range(len(items)):
        store_ids.append(idx)
        print(store_ids)
        data.append((idx, items[idx][0], items[idx][1], items[idx][2], items[idx][3], items[idx][4], items[idx][5], items[idx][6], items[idx][7], items[idx][8], items[idx][9]))
    # print(data)

    #sorted_high_data, sorted_low_data = sort_price(text1, data)
    MyApp2().run()
    # df = pd.read_excel('coat.xlsx')
    # Stores the data





#run_data("gaming laptop", False)

'''
data = {"Website": "Amazon",
        "Description": descrpition,
        "Price": price,
        "Stock": instock,
        "Rating": rating,
        "Review": review,
        "Color": colors,
        "Size": all_size,
        "Link": link
        }
        '''


#Use global data varriable
#Covert the data to high and low again
#Display it